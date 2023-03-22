import  openpyxl  as  op  #openpyxl 모듈 import

wb = op.load_workbook(r"test.xlsx") #워크북 객체 생성(파일명 : test.xlsx)
ws = wb.active  #활성화 되어있는 시트 설정(시트명 : "업")

#방법 1 : Sheet의 Cell 속성 사용하기
data1 = ws.cell(row=1, column=2).value

#방법 2 : 엑셀 인덱스(Range) 사용하기
data2 = ws["B1"].value

#위 결과 출력해보기
print("cell(1,2) : ", data1)
print('Range("B1"):', data2)